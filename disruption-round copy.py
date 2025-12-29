import pandas as pd
from openpyxl.styles import Font, PatternFill

# Create the repeating pattern 4,3,2,1 for numbers 1 to 50
numbers = list(range(1, 51))
pattern = [4, 3, 2, 1]

# Create DataFrame
data = []
for i, num in enumerate(numbers):
    pattern_value = pattern[i % 4]
    rotation = "C" if pattern_value <= 2 else ("B" if pattern_value == 3 else "A")
    next_c = 4 - (i % 4) if pattern_value != 1 else 0
    next_c_text = f"In {next_c}" if pattern_value != 1 else "In 1"

    data.append(
        {
            "Round": num,
            "Pattern": pattern_value,
            "Rotation": rotation,
            "Next C Rotation": next_c_text,
        }
    )

df = pd.DataFrame(data)

# Display as a formatted table
print("WARFRAME DISRUPTION ROTATION TABLE")
print("=" * 50)
print(df.to_string(index=False))

# Option 1: Save to CSV
df.to_csv("warframe_disruption_rotations.csv", index=False)
print("\n✓ Saved to 'warframe_disruption_rotations.csv'")

# Option 2: Save to Excel with formatting
try:
    with pd.ExcelWriter(
        "warframe_disruption_rotations.xlsx", engine="openpyxl"
    ) as writer:
        df.to_excel(writer, sheet_name="Rotations", index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets["Rotations"]

        # Add some basic formatting

        # Format header
        header_fill = PatternFill(
            start_color="4CAF50", end_color="4CAF50", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Format C rotations
        for row in range(2, len(df) + 2):
            if worksheet[f"C{row}"].value == "C":  # Rotation column
                worksheet[f"C{row}"].fill = PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                )
                worksheet[f"C{row}"].font = Font(bold=True)

    print("✓ Saved to 'warframe_disruption_rotations.xlsx'")
except ImportError:
    print("\nNote: Install openpyxl for Excel support: pip install openpyxl")
